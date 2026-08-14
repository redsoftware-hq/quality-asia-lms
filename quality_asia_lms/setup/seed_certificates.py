"""Seed batches of Internal Auditor training certificates from client spreadsheets.

The client ships a spreadsheet of completed trainings; each row becomes a
published `LMS Certificate` plus the `User` / `LMS Enrollment` needed to hang it
off. Two batches are defined in BATCHES below; each is driven by a patch.

The spreadsheets hold candidate names, emails and phone numbers, so they are NOT
committed to this repo. Place the file on the site and run:

    scp "<xlsx>" <server>:/.../sites/<site>/private/files/
    bench --site <site> execute quality_asia_lms.setup.seed_certificates.reconcile
    bench --site <site> execute quality_asia_lms.setup.seed_certificates.run
    # then delete the xlsx from the server

Both entry points take a `batch` kwarg naming a key in BATCHES; it defaults to
the 11-08 batch, e.g. --kwargs "{'batch': 'itsms'}".

`reconcile()` is read-only — run it first. `run()` is idempotent: it skips rows
whose (member, course) certificate already exists, so a partial run resumes
cleanly.

Certificate numbering
---------------------
Numbers are assigned by *sheet row index*, not by autoname and not by a running
counter:

    row i (0-based, blank rows dropped)  ->  IAC-{start + i:05d}

This makes the row->number mapping identical on every run. A running counter
would drift on a re-run, because skipped rows would no longer consume a number.

`start` is detected from the site's current highest IAC number on the first run
and then stored under the batch's own key, so it survives certificates issued in
the meantime by the live portal without ever moving underneath a resumed row.
See `_resolve_start`.

Batches run back to back in one `bench migrate`, in patches.txt order. Each one
claims the top of its block on its first insert, so the next batch's
`_scan_max_iac` sees that ceiling and allocates cleanly above it.

A row that is skipped, unmapped or failed leaves its number unused rather than
letting later rows slide down into it — that is what keeps the mapping stable,
at the cost of gaps in the block when something goes wrong. Fixing the cause and
re-running gives those rows their original numbers back.

Emails are suppressed entirely: `is_migrated=1` short-circuits
`QALMSCertificate.send_certification_email`, so no Email Queue rows and no PDF
renders happen during the run.
"""

import os
import re
from dataclasses import dataclass, field

import frappe
from frappe.utils import getdate

from quality_asia_lms.overrides import dwm_migration as dwm

TEMPLATE = "QA Certificate"

# Matches the tolerant pattern QALMSCertificate._next_iac_number uses, so we
# agree with autoname about what counts as a number — including the legacy
# "IAC -XXXXX" space variants from the DWM import.
_IAC_RE = re.compile(r"^IAC\s*-\s*0*(\d+)$", re.IGNORECASE)

# Users created for rows whose email cell holds something that isn't an address.
PLACEHOLDER_DOMAIN = "@placeholder.qualityasia.in"

# Column order in the sheets, left to right. Both batches share this layout.
COL_CANDIDATE = 1
COL_PROGRAM = 2
COL_TRAINING_DATES = 3
COL_ISSUE_DATE = 4
COL_EMAIL = 5
COL_CONTACT = 6

# The ITSMS course does not exist on the site and has no content — it exists only
# to anchor its certificates, so it is unpublished and certification-free.
#
# The title is deliberately the dangling phrase "ITSMS Awareness and". The
# QA Certificate print format renders
#     HAS SUCESSFULLY COMPLETED {{ course_title }} INTERNAL AUDITOR TRAINING PROGRAM
# so this title makes the sentence read "ITSMS Awareness and Internal Auditor
# Training Program" — the client's required wording — with no change to the
# shared print format, which every other certificate on the site also renders
# through. The ISO standard lives in the slug, short_introduction and tags
# instead of the title.
ITSMS_COURSE = {
	"name": "iso-20000-1-2018-itsms-awareness-and-internal-auditor-training",
	"title": "ITSMS Awareness and",
	"short_introduction": "ISO 20000-1:2018 - IT Service Management System (ITSMS)",
	"description": "ISO 20000-1:2018 ITSMS Awareness and Internal Auditor Training.",
	"tags": "ISO 20000-1:2018",
	"published": 0,
	"enable_certification": 0,
}


@dataclass(frozen=True)
class Batch:
	"""One spreadsheet's worth of certificates."""

	key: str
	data_file: str
	sheet: str
	expected_rows: int
	start_key: str  # tabDefaultValue key holding this batch's allocated block start
	report_prefix: str
	# Program text -> LMS Course name, for batches whose sheet wording differs
	# from the course title. Empty means resolve by title as usual.
	course_by_program: dict = field(default_factory=dict)
	# Course to create before seeding, when the batch's course is not on the site.
	course_spec: dict | None = None


BATCHES = {
	"11_08": Batch(
		key="11_08",
		data_file="Internal Auditor Training Certificate - 11-08.xlsx",
		sheet="Internal Auditor Training Cert",
		expected_rows=611,
		start_key="ia_cert_seed_11_08_start",
		report_prefix="ia_cert_seed_11_08",
	),
	"itsms": Batch(
		key="itsms",
		data_file="Internal Auditor Training Certificate - 12-08 - Nelito.xlsx",
		sheet="Internal Auditor Training Certi",
		expected_rows=13,
		start_key="ia_cert_seed_12_08_itsms_start",
		report_prefix="ia_cert_seed_12_08_itsms",
		course_by_program={
			"itsms awareness and internal auditor training": ITSMS_COURSE["name"],
		},
		course_spec=ITSMS_COURSE,
	),
}

DEFAULT_BATCH = "11_08"


def _batch(key):
	batch = BATCHES.get(key)
	if not batch:
		frappe.throw(f"Unknown batch {key!r}. Known: {sorted(BATCHES)}")
	return batch


def _default_path(batch):
	return frappe.get_site_path("private", "files", batch.data_file)


def _log(msg):
	print(f"[qa-cert-seed] {msg}")


def _rows(path, batch):
	"""Read the sheet into dicts, dropping fully-blank rows.

	Order is the sheet's own order and is what the numbering indexes against, so
	it must stay deterministic.
	"""
	from openpyxl import load_workbook

	wb = load_workbook(path, data_only=True, read_only=True)
	if batch.sheet not in wb.sheetnames:
		frappe.throw(f"Sheet {batch.sheet!r} not found in {path} (found: {wb.sheetnames})")

	rows = []
	for raw in wb[batch.sheet].iter_rows(min_row=2, values_only=True):
		if not any(raw):
			continue
		contact = raw[COL_CONTACT]
		rows.append(
			{
				"candidate_name": (raw[COL_CANDIDATE] or "").strip(),
				"program": (raw[COL_PROGRAM] or "").strip(),
				"training_dates": (raw[COL_TRAINING_DATES] or "").strip(),
				"issue_date": getdate(raw[COL_ISSUE_DATE]) if raw[COL_ISSUE_DATE] else None,
				"email": (raw[COL_EMAIL] or "").strip().lower(),
				"contact": str(contact).strip() if contact is not None else "",
			}
		)
	wb.close()
	return rows


def _usable_email(email):
	"""True only for an address we can actually create an account on."""
	email = (email or "").strip()
	return bool(email) and bool(dwm.EMAIL_RE.match(email)) and dwm._is_valid_email(email)


def _course_for(program, batch):
	"""Resolve a sheet's program text to an LMS Course, honouring batch overrides."""
	override = batch.course_by_program.get((program or "").strip().lower())
	return override or dwm._resolve_course(program)


def _ensure_course(spec):
	"""Create the batch's course if absent. Idempotent; never edits an existing one.

	`instructors` is a required field, filled by the auto_assign_mentor
	before_insert hook (hooks.py) — ignore_mandatory covers sites where the
	mentor user does not exist.
	"""
	if frappe.db.exists("LMS Course", spec["name"]):
		return spec["name"]
	doc = frappe.get_doc({**spec, "doctype": "LMS Course"})
	doc.flags.ignore_permissions = True
	doc.flags.ignore_mandatory = True
	doc.insert(set_name=spec["name"])
	_log(f"created course {spec['name']} (title={spec['title']!r}, published={spec['published']})")
	return doc.name


def _cert_name(index, start):
	return f"IAC-{start + index:05d}"


def _scan_max_iac():
	"""Highest IAC number currently on the site, by the same rules as autoname."""
	top = 0
	for name in frappe.get_all("LMS Certificate", pluck="name"):
		match = _IAC_RE.match(name or "")
		if match:
			top = max(top, int(match.group(1)))
	return top


def _resolve_start(row_count, batch, persist=False):
	"""First number of this batch's block: detected once, then reused forever.

	This is a live portal — real users earn certificates through
	QALMSCertificate.autoname between the day a batch is scoped and the day it is
	deployed, so a hardcoded start would collide and abort the run.

	But the value must also stay put across re-runs. If it were recomputed each
	time, a resumed run would see a higher max and give row N a different number
	than its first attempt did — leaving gaps and an unpredictable mapping. So
	the first run stores its allocation and every later run reads it back, which
	is also what lets a FAILED row reclaim its original number on a retry.
	"""
	stored = frappe.db.get_default(batch.start_key)
	if stored:
		return int(stored)

	start = _scan_max_iac() + 1
	names = [_cert_name(i, start) for i in range(row_count)]
	clash = frappe.get_all("LMS Certificate", filters={"name": ["in", names]}, pluck="name")
	if clash:
		# Can't happen while the block sits above the detected max — a hard stop
		# beats minting a duplicate number.
		frappe.throw(f"Cannot allocate a free block at {start}: {clash[:5]} already exist.")

	if persist:
		frappe.db.set_default(batch.start_key, start)
	return start


def _classify_range(rows, start, batch):
	"""Split occupied target numbers into (ours, foreign).

	A target number is *ours* when the certificate sitting on it already has the
	same member and course as the row that maps to it — that is a previous run of
	this same batch, and resuming over it is correct.

	Anything else is *foreign*: the number belongs to an unrelated certificate,
	which means the allocated block start is wrong and we must not write at all.

	Distinguishing the two is what lets the pre-flight guard coexist with the
	idempotent resume. A blanket "is this number taken?" check would abort every
	resume, since a partial run's own rows occupy their own numbers.
	"""
	names = [_cert_name(i, start) for i in range(len(rows))]
	existing = {
		row["name"]: row
		for row in frappe.get_all(
			"LMS Certificate", filters={"name": ["in", names]}, fields=["name", "member", "course"]
		)
	}

	ours, foreign = [], []
	for index, row in enumerate(rows):
		found = existing.get(_cert_name(index, start))
		if not found:
			continue
		course = _course_for(row["program"], batch)
		same_member = not row["email"] or found["member"] == row["email"]
		if found["course"] == course and same_member:
			ours.append(found["name"])
		else:
			foreign.append(found["name"])
	return ours, foreign


# ── entry points ─────────────────────────────────────────────────────────────


def reconcile(path=None, batch=DEFAULT_BATCH):
	"""Read-only preview: course mapping, numbering block, and expected skips.

	Writes nothing. Run this before run() and check every line.
	"""
	batch = _batch(batch)
	path = path or _default_path(batch)
	if not os.path.exists(path):
		frappe.throw(f"Data file not found: {path}")

	dwm._course_titles = None  # force a fresh course cache
	rows = _rows(path, batch)

	_log(f"batch: {batch.key}")
	_log(f"file:  {path}")
	_log(f"rows:  {len(rows)} (expected {batch.expected_rows})")
	if len(rows) != batch.expected_rows:
		_log("!! ROW COUNT MISMATCH — wrong file, or expected_rows needs updating")

	if batch.course_spec and not frappe.db.exists("LMS Course", batch.course_spec["name"]):
		_log(f"course {batch.course_spec['name']} does not exist yet — run() will create it")

	programs = {}
	for row in rows:
		programs[row["program"]] = programs.get(row["program"], 0) + 1

	_log("")
	_log(f"{'rows':>6}  {'program':<46} resolution")
	unmapped = 0
	for program, count in sorted(programs.items(), key=lambda kv: -kv[1]):
		course = _course_for(program, batch)
		if not course:
			unmapped += count
		_log(f"{count:>6}  {program:<46} {'-> ' + course if course else 'UNMAPPED (rows will be skipped)'}")

	_log("")
	start = _resolve_start(len(rows), batch, persist=False)  # preview only — never stores
	stored = frappe.db.get_default(batch.start_key)
	_log(f"block start: {start} ({'already allocated' if stored else 'would be allocated now'})")
	if rows:
		_log(f"numbering: {_cert_name(0, start)} .. {_cert_name(len(rows) - 1, start)}")
	ours, foreign = _classify_range(rows, start, batch)
	_log(f"already seeded by a previous run (will resume): {len(ours)}")
	_log(f"FOREIGN collisions (would abort): {len(foreign)}{' -> ' + str(foreign[:20]) if foreign else ''}")

	existing = 0
	for row in rows:
		course = _course_for(row["program"], batch)
		if course and frappe.db.exists(
			"LMS Certificate", {"member": row["email"], "course": course}
		):
			existing += 1

	bad_email = [r["candidate_name"] for r in rows if not _usable_email(r["email"])]
	_log("")
	_log(f"already certified (will skip): {existing}")
	_log(f"unusable emails (get placeholder accounts): {len(bad_email)}{' -> ' + str(bad_email[:10]) if bad_email else ''}")
	_log(f"unmapped rows (will skip): {unmapped}")
	_log(f"=> expected created: {len(rows) - existing - unmapped}")
	return (
		f"batch={batch.key} rows={len(rows)} resumable={len(ours)} foreign={len(foreign)} "
		f"existing={existing} unmapped={unmapped}"
	)


def run_from_patch(batch_key):
	"""Entry point for the seeding patches — the only thing they should call.

	Both patches route through here so a signature change in this module cannot
	leave one of them calling a stale API. That is exactly how the 11-08 patch
	broke on its first deploy: it still called _default_path() with no argument
	after the batch refactor added one, raised TypeError before reaching its own
	file check, and was skipped by `bench migrate --skip-failing`.

	Returns None when the spreadsheet is absent, having logged to Error Log —
	the patch is then recorded as executed and will not retry on its own.
	"""
	batch = _batch(batch_key)
	path = _default_path(batch)
	if not os.path.exists(path):
		frappe.log_error(
			message=(
				f"Expected the spreadsheet for batch {batch_key!r} at {path}, but it "
				f"was not there.\n"
				f"This patch is now recorded as executed and will NOT re-run.\n"
				f"Upload the file, then run:\n"
				f"  bench --site {frappe.local.site} execute "
				f"quality_asia_lms.setup.seed_certificates.run "
				f"--kwargs \"{{'batch': '{batch_key}'}}\""
			),
			title=f"cert seed {batch_key}: data file missing, patch consumed",
		)
		return None
	return run(path, batch=batch_key)


def run(path=None, limit=None, batch=DEFAULT_BATCH):
	"""Create the certificates. Idempotent on (member, course)."""
	batch = _batch(batch)
	path = path or _default_path(batch)
	if not os.path.exists(path):
		frappe.throw(f"Data file not found: {path}")

	# reset the shared caches in dwm_migration so we see current DB state
	dwm._course_titles = None
	dwm._user_cache = set()

	rows = _rows(path, batch)
	if len(rows) != batch.expected_rows:
		frappe.throw(
			f"Refusing to run: sheet has {len(rows)} rows, expected {batch.expected_rows}. "
			f"Wrong file, or update the batch's expected_rows deliberately."
		)

	if batch.course_spec:
		_ensure_course(batch.course_spec)

	if limit:
		rows = rows[: int(limit)]

	start = _resolve_start(len(rows), batch, persist=True)
	_log(f"batch {batch.key} block start: {start} -> {_cert_name(0, start)} .. {_cert_name(len(rows) - 1, start)}")

	# Pre-flight: no target number may belong to an unrelated certificate, or we
	# write nothing at all. Numbers held by this batch's own earlier run are fine
	# — those rows get skipped below and the run resumes.
	ours, foreign = _classify_range(rows, start, batch)
	if foreign:
		frappe.throw(
			f"Refusing to run: {len(foreign)} target certificate numbers belong to "
			f"unrelated certificates (e.g. {foreign[:5]}). The allocated block start "
			f"({start}, stored under default {batch.start_key!r}) is wrong for this site. "
			f"Do not shift numbers ad hoc — investigate, then clear that default to "
			f"reallocate."
		)
	if ours:
		_log(f"resuming: {len(ours)} certificate(s) from a previous run already in place")

	placeholder_log: list = []
	report: list = []
	created = skipped = unmapped = failed = 0

	# bypass the user-creation throttle; also makes send_certification_email a no-op
	prev_import, prev_install = frappe.flags.in_import, frappe.flags.in_install
	frappe.flags.in_import = True
	frappe.flags.in_install = True

	# Highest index first, so the top of the block is claimed on the very first
	# insert. Reserving numbers in tabDefaultValue means nothing to autoname —
	# it mints max+1 — so until the top exists, a certificate earned on the live
	# portal lands inside our block and blocks the resume as a foreign collision.
	# Claiming the top makes autoname skip past the whole range. Numbering is
	# index-based, so processing order does not affect which number a row gets.
	order = list(range(len(rows)))
	if len(order) > 1:
		order = [order[-1], *order[:-1]]

	try:
		for index in order:
			row = rows[index]
			cert_name = _cert_name(index, start)
			course = _course_for(row["program"], batch)
			if not course:
				unmapped += 1
				report.append((cert_name, row["email"], row["program"], "UNMAPPED"))
				continue
			try:
				member = dwm._user(
					row["candidate_name"], row["email"], row["contact"], placeholder_log
				)

				if frappe.db.exists("LMS Certificate", {"member": member, "course": course}):
					skipped += 1
					report.append((cert_name, member, row["program"], "SKIPPED (already certified)"))
					continue

				enrollment = frappe.db.get_value(
					"LMS Enrollment", {"member": member, "course": course}, ["name", "progress"], as_dict=True
				)
				if not enrollment:
					doc = frappe.get_doc(
						{
							"doctype": "LMS Enrollment",
							"member": member,
							"course": course,
							"progress": 100,
						}
					)
					doc.flags.ignore_permissions = True
					doc.flags.ignore_mandatory = True
					doc.insert()
					enrollment_name = doc.name
				else:
					enrollment_name = enrollment.name
					# They completed the training offline; a stale sub-100 progress would
					# fail LMSCertificate.validate_course_enrollment.
					if (enrollment.progress or 0) < 100:
						frappe.db.set_value(
							"LMS Enrollment", enrollment_name, "progress", 100, update_modified=False
						)

				cert = frappe.get_doc(
					{
						"doctype": "LMS Certificate",
						"member": member,
						"course": course,
						"issue_date": row["issue_date"],
						"template": TEMPLATE,
						"published": 1,
						"training_dates": row["training_dates"],
						"candidate_name_as_printed": row["candidate_name"],
						"is_migrated": 1,
					}
				)
				cert.flags.ignore_permissions = True
				cert.insert(set_name=cert_name)  # bypasses QALMSCertificate.autoname

				frappe.db.set_value(
					"LMS Enrollment", enrollment_name, "certificate", cert_name, update_modified=False
				)

				created += 1
				placeholder = member.endswith(PLACEHOLDER_DOMAIN)
				report.append(
					(cert_name, member, row["program"], "CREATED (placeholder email)" if placeholder else "CREATED")
				)
				if created % 100 == 0:
					frappe.db.commit()
					_log(f"{created} created")
			except Exception:
				frappe.log_error(title=f"IA cert seed failed: {cert_name}")
				failed += 1
				report.append((cert_name, row["email"], row["program"], "FAILED"))

		frappe.db.commit()
	finally:
		frappe.flags.in_import = prev_import
		frappe.flags.in_install = prev_install

	_write_reports(report, placeholder_log, batch)

	summary = (
		f"batch={batch.key} created={created} skipped={skipped} unmapped={unmapped} "
		f"failed={failed} placeholders={len(placeholder_log)}"
	)
	_log(summary)
	return summary


def _write_reports(report, placeholder_log, batch):
	"""Save the row->number mapping (and any placeholder users) as private Files."""
	import csv
	import io

	buf = io.StringIO()
	writer = csv.writer(buf)
	writer.writerow(("certificate", "member", "program", "status"))
	writer.writerows(sorted(report))  # run order is top-first; report reads better sorted
	dwm._save_private_file(f"{batch.report_prefix}_mapping.csv", buf.getvalue())

	if placeholder_log:
		buf = io.StringIO()
		csv.writer(buf).writerows(
			[("candidate_name", "mobile", "placeholder_email"), *placeholder_log]
		)
		dwm._save_private_file(f"{batch.report_prefix}_placeholders.csv", buf.getvalue())

	frappe.db.commit()
